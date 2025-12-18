from pathlib import Path
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def _autosize_columns(ws):
    """
    Autosize Excel columns based on content length.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)


def _freeze_panes(ws):
    """
    Freeze header row and first column.
    """
    ws.freeze_panes = "B2"


def _bold_totals(ws):
    """
    Bold the Total row and Total Result column.
    """
    for cell in ws[1]:
        if str(cell.value).lower() == "total result":
            total_col = cell.column
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=total_col).font = Font(bold=True)

    for r in range(1, ws.max_row + 1):
        if str(ws.cell(row=r, column=1).value).lower() == "total":
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).font = Font(bold=True)


def write_excel_report(
    matrices: dict,
    report_date,
    output_dir: Path
):
    """
    Write full-detail Excel report:
    - Actuals
    - YoY Variance
    - Budget Variance
    """

    output_dir.mkdir(parents=True, exist_ok=True)

    file_path = output_dir / f"Daily_Radiology_Report_{report_date}.xlsx"

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        sheet_map = {
            "Actuals": matrices["actual"],
            "YoY Variance": matrices["yoy"],
            "Budget Variance": matrices["budget"],
        }

        for sheet_name, df in sheet_map.items():
            df_out = df.copy().fillna(0)
            df_out.to_excel(
                writer,
                sheet_name=sheet_name,
                startrow=0,
                startcol=0
            )

            ws = writer.book[sheet_name]

            _freeze_panes(ws)
            _autosize_columns(ws)
            _bold_totals(ws)

    return file_path
